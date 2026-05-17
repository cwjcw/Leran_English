from tempfile import NamedTemporaryFile
from pathlib import Path

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile, status
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.api.deps import get_current_user
from app.db.models import User
from app.db.models import Word
from app.db.session import get_db
from app.schemas import BulkWordImport, WordCreate, WordRead, WordUpdate
from app.services.ai_service import AIService


router = APIRouter(prefix="/words", tags=["words"])


def _row_value(row: dict, *names: str) -> str | None:
    for name in names:
        value = row.get(name)
        if value is not None and str(value).strip():
            return str(value).strip()
    return None


@router.get("", response_model=list[WordRead])
def list_words(
    tag: str | None = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> list[Word]:
    stmt = select(Word).order_by(Word.id)
    if tag:
        stmt = stmt.where(Word.dynamic_tags.like(f"%{tag}%"))
    return list(db.scalars(stmt).all())


@router.post("", response_model=WordRead, status_code=status.HTTP_201_CREATED)
def create_word(
    payload: WordCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> Word:
    word = Word(**payload.model_dump())
    db.add(word)
    try:
        db.commit()
    except Exception as exc:
        db.rollback()
        raise HTTPException(status_code=409, detail="Word already exists or payload is invalid.") from exc
    db.refresh(word)
    return word


@router.post("/bulk", response_model=list[WordRead], status_code=status.HTTP_201_CREATED)
async def bulk_import_words(
    payload: BulkWordImport,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> list[Word]:
    created: list[Word] = []
    items = [item.model_copy() for item in payload.words]
    missing_translation = [item.word for item in items if not item.translation]
    if missing_translation:
        try:
            enriched = await AIService().enrich_words(missing_translation)
            for item in items:
                data = enriched.get(item.word.lower())
                if data:
                    if not data.get("is_valid", True):
                        reason = data.get("reason") or "不是有效英文单词。"
                        raise HTTPException(status_code=422, detail=f"{item.word}: {reason}")
                    item.translation = data.get("translation", item.translation)
                    item.phonetic = item.phonetic or data.get("phonetic")
        except HTTPException:
            raise
        except Exception as exc:
            raise HTTPException(status_code=502, detail=f"Unable to validate words: {exc}") from exc

    for item in items:
        existing = db.scalar(select(Word).where(Word.word == item.word))
        if existing:
            for key, value in item.model_dump().items():
                setattr(existing, key, value)
            created.append(existing)
        else:
            word = Word(**item.model_dump())
            db.add(word)
            created.append(word)
    db.commit()
    for word in created:
        db.refresh(word)
    return created


@router.post("/excel", response_model=list[WordRead], status_code=status.HTTP_201_CREATED)
async def import_words_from_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> list[Word]:
    suffix = Path(file.filename or "words.xlsx").suffix
    if suffix.lower() not in {".xlsx", ".xlsm"}:
        raise HTTPException(status_code=422, detail="Please upload an .xlsx file.")

    content = await file.read()
    with NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
        tmp.write(content)
        tmp_path = Path(tmp.name)
    try:
        workbook = load_workbook(tmp_path, read_only=True, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
    finally:
        if "workbook" in locals():
            workbook.close()
        tmp_path.unlink(missing_ok=True)

    if not rows:
        raise HTTPException(status_code=422, detail="Excel file is empty.")

    headers = [str(cell or "").strip().lower() for cell in rows[0]]
    words: list[WordCreate] = []
    for values in rows[1:]:
        row = {headers[index]: values[index] for index in range(min(len(headers), len(values)))}
        word = _row_value(row, "word", "单词", "英文")
        if not word:
            continue
        words.append(
            WordCreate(
                word=word,
                translation=_row_value(row, "translation", "中文", "释义") or "",
                phonetic=_row_value(row, "phonetic", "音标"),
                dynamic_tags=_row_value(row, "tag", "tags", "标签") or f"wordbook-{current_user.id}",
                textbook=_row_value(row, "textbook", "教材"),
                grade=_row_value(row, "grade", "年级"),
                unit=_row_value(row, "unit"),
                lesson=_row_value(row, "lesson"),
            )
        )
    if not words:
        raise HTTPException(status_code=422, detail="No words found in Excel file.")
    return await bulk_import_words(BulkWordImport(words=words), db, current_user)


@router.get("/{word_id}", response_model=WordRead)
def get_word(
    word_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> Word:
    word = db.get(Word, word_id)
    if not word:
        raise HTTPException(status_code=404, detail="Word not found.")
    return word


@router.patch("/{word_id}", response_model=WordRead)
def update_word(
    word_id: int,
    payload: WordUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> Word:
    word = db.get(Word, word_id)
    if not word:
        raise HTTPException(status_code=404, detail="Word not found.")
    for key, value in payload.model_dump(exclude_unset=True).items():
        setattr(word, key, value)
    db.commit()
    db.refresh(word)
    return word


@router.delete("/{word_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_word(
    word_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> None:
    word = db.get(Word, word_id)
    if not word:
        raise HTTPException(status_code=404, detail="Word not found.")
    db.delete(word)
    db.commit()
